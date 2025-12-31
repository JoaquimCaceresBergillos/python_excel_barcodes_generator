import time
import os
import math
import shutil
import pandas as pd
from barcode import Code128, EAN13
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

inicio = time.time()  # guardar tiempo inicial

# ---------- CONFIGURACIÓN ----------
ARCHIVO_ENTRADA = "articulos.xlsx"  # Archivo original
CARPETA_IMAGENES = "barcodes"       # carpeta para guardar las imágenes de los códigos de barras



COLUMNA_CODIGO = "cod_barras"  # columna con el número original

FILAS_POR_ARCHIVO = 5000  # dividir Excel cada N filas


GENERAR_CODE128 = False
OPTIONS_CODE128 = {
    "module_width": 0.2,    # ancho de cada barra
    "module_height": 8,    # altura de las barras
    "font_size": 8,        # tamaño del número debajo
    "text_distance": 3.5,   # distancia entre barra y texto
    "quiet_zone": 4,        # margen en blanco a los lados
    "background": "white",  # color de fondo
    "foreground": "black",  # color de las barras
    "write_text": True,     # mostrar número debajo
    "dpi": 200              # resolución
}
COLUMNA_CODE128 = "code128"  # columna donde se generará el code128
OPTIONS_CODE128_ANCHO_IMAGEN = 150 # Ancho de la imagen del code128 en Excel
OPTIONS_CODE128_ALTO_IMAGEN = 45 # Alto de la imagen del code128 en Excel



GENERAR_EAN13 = True
OPTIONS_EAN13 = {
    "module_width": 0.2,    # ancho de cada barra
    "module_height": 8,    # altura de las barras
    "font_size": 6,        # tamaño del número debajo
    "text_distance": 2.5,   # distancia entre barra y texto
    "quiet_zone": 4,        # margen en blanco a los lados
    "background": "white",  # color de fondo
    "foreground": "black",  # color de las barras
    "write_text": True,     # mostrar número debajo
    "dpi": 200              # resolución
}
COLUMNA_EAN13 = "ean13"      # columna donde se generará el EAN13
OPTIONS_EAN13_ANCHO_IMAGEN = 150  # Ancho de la imagen del ean13 en Excel
OPTIONS_EAN13_ALTO_IMAGEN = 45 # Alto de la imagen del ean13 en Excel



IS_WINDOWS = os.name == "nt" # Determinar si la terminal es Windows
EMOJI_OK = "✅" if not IS_WINDOWS else "[OK]"
EMOJI_CLOCK = "⏳" if not IS_WINDOWS else "[TIEMPO]"
EMOJI_INFO = "ℹ️ " if not IS_WINDOWS else "[INFO]"



# -------------------------------------------------------------------------------------------------
print(f" ")
print(f" ")
print(f" ")
print(f"******************************************************")
print(f"*** GENERADOR DE CÓDIGOS DE BARRAS PARA EXCEL v1.0 ***")
print(f"******************** By JoaquimCB ********************")
print(f"******************************************************")
print(f" ")
print(f" ")

# ---------- CONFIGURACIÓN INTERACTIVA ----------
while True:
    ARCHIVO_ENTRADA = input("Ingrese el nombre del archivo Excel de entrada (con extensión .xlsx): ").strip()
    if os.path.exists(ARCHIVO_ENTRADA) and ARCHIVO_ENTRADA.endswith(".xlsx"):
        break
    else:
        print("❌ Archivo no encontrado o extensión incorrecta. Intente de nuevo.")

"""
while True:
    ARCHIVO_ENTRADA = input("Ingrese el nombre del archivo Excel de entrada (con extensión .xlsx): ").strip()
    if os.path.exists(ARCHIVO_ENTRADA) and ARCHIVO_ENTRADA.endswith(".xlsx"):
        break
    else:
        print("❌ Archivo no encontrado o extensión incorrecta. Intente de nuevo.")

while True:
    COLUMNA_CODIGO = input("Ingrese el nombre de la columna que contiene los códigos de barras: ").strip()
    if COLUMNA_CODIGO:
        break
    else:
        print("❌ Debe ingresar un nombre de columna válido.")

while True:
    try:
        FILAS_POR_ARCHIVO = int(input("Ingrese la cantidad de filas por archivo: ").strip())
        if FILAS_POR_ARCHIVO > 0:
            break
        else:
            print("❌ Debe ser un número mayor a 0.")
    except ValueError:
        print("❌ Ingrese un número válido.")
        
print(f" ")
print(f" ")
"""

print(f"{EMOJI_INFO} Ejecutando programa...")
print(f" ")
print(f" ")

# Borra el contenido de las carpetas si existen
if os.path.exists(CARPETA_IMAGENES):
    shutil.rmtree(CARPETA_IMAGENES)

# Crear carpeta principal de imágenes si no existe
os.makedirs(CARPETA_IMAGENES, exist_ok=True)

# Leer Excel completo
df = pd.read_excel(ARCHIVO_ENTRADA)

# Dividir en bloques
total_filas = len(df)
num_archivos = (total_filas // FILAS_POR_ARCHIVO) + (1 if total_filas % FILAS_POR_ARCHIVO else 0)

for bloque in range(num_archivos):
    inicio_bloque = bloque * FILAS_POR_ARCHIVO
    fin_bloque = inicio_bloque + FILAS_POR_ARCHIVO
    df_bloque = df.iloc[inicio_bloque:fin_bloque].copy()
    
    # Crear carpetas de imágenes para este bloque
    bloque_folder = f"{CARPETA_IMAGENES}/bloque_{bloque+1}"
    os.makedirs(f"{bloque_folder}/code128", exist_ok=True)
    os.makedirs(f"{bloque_folder}/ean13", exist_ok=True)
    
    # Columnas vacías
    if GENERAR_CODE128:
        df_bloque[COLUMNA_CODE128] = ""
    if GENERAR_EAN13:
        df_bloque[COLUMNA_EAN13] = ""
    
    rutas_code128 = []
    rutas_ean13  = []
    
    # ---------- Generación de imágenes ----------
    print(f"{EMOJI_INFO} Generando imágenes de códigos de barras para bloque {bloque+1} ...")
    for i, codigo in enumerate(df_bloque[COLUMNA_CODIGO]):
        # Saltar valores vacíos
        if codigo is None or (isinstance(codigo, float) and math.isnan(codigo)):
            rutas_code128.append(None)
            rutas_ean13.append(None)
            continue

        codigo_str = str(int(codigo))
        
        # ---------- Code128 ----------
        if GENERAR_CODE128:
            ruta_c128 = f"{bloque_folder}/code128/{codigo_str}_code128_{i}.png"
            Code128(codigo_str, writer=ImageWriter()).save(ruta_c128.replace(".png",""), options=OPTIONS_CODE128)
            rutas_code128.append(ruta_c128)
        
        # ---------- EAN13 ----------
        if GENERAR_EAN13:
            ean13_str = codigo_str.zfill(13)
            ruta_ean13 = f"{bloque_folder}/ean13/{ean13_str}_ean13_{i}.png"
            EAN13(ean13_str, writer=ImageWriter()).save(ruta_ean13.replace(".png",""), options=OPTIONS_EAN13)
            rutas_ean13.append(ruta_ean13)
    print(f"{EMOJI_OK} Imágenes generadas para bloque {bloque+1}")

    # Guardar Excel temporal
    temp_excel = f"{bloque_folder}/temp.xlsx"
    df_bloque.to_excel(temp_excel, index=False)
    
    # Abrir Excel con openpyxl
    wb = load_workbook(temp_excel)
    ws = wb.active

    # Función para encontrar índice de columna por nombre
    def indice_columna(nombre_columna):
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == nombre_columna:
                return idx
        return None

    idx_c128 = indice_columna(COLUMNA_CODE128) if GENERAR_CODE128 else None
    idx_ean13 = indice_columna(COLUMNA_EAN13) if GENERAR_EAN13 else None


    # ---------- Inserción de imágenes ----------
    print(f"{EMOJI_INFO} Insertando imágenes en el Excel del bloque {bloque+1} ...")
    for i in range(2, ws.max_row + 1):
        if GENERAR_CODE128 and rutas_code128[i-2]:
            img_c128 = Image(rutas_code128[i-2])
            img_c128.width = OPTIONS_CODE128_ANCHO_IMAGEN
            img_c128.height = OPTIONS_CODE128_ALTO_IMAGEN
            celda_c128 = f"{ws.cell(row=1, column=idx_c128).column_letter}{i}"
            ws.add_image(img_c128, celda_c128)
        if GENERAR_EAN13 and rutas_ean13[i-2]:
            img_ean13 = Image(rutas_ean13[i-2])
            img_ean13.width = OPTIONS_EAN13_ANCHO_IMAGEN
            img_ean13.height = OPTIONS_EAN13_ALTO_IMAGEN
            celda_ean13 = f"{ws.cell(row=1, column=idx_ean13).column_letter}{i}"
            ws.add_image(img_ean13, celda_ean13)
        ws.row_dimensions[i].height = 50
    print(f"{EMOJI_OK} Imágenes insertadas en el Excel del bloque {bloque+1}")

     # ---------- Guardar Excel final ----------
    print(f"{EMOJI_INFO} Generando Excel del bloque {bloque+1} ...")
    NOMBRE_BASE_ENTRADA = os.path.splitext(os.path.basename(ARCHIVO_ENTRADA))[0]
    archivo_salida = f"{NOMBRE_BASE_ENTRADA}_barcodes_{bloque+1}.xlsx"
    wb.save(archivo_salida)
    os.remove(temp_excel)

    print(f"{EMOJI_OK} Generado {archivo_salida} con {len(df_bloque)} filas")
    print(f" ")
    print(f" ")

fin = time.time()
tiempo_total = fin - inicio
minutos = int(tiempo_total // 60)
segundos = int(tiempo_total % 60)
print("--------------------------------")
print(f" ")
print(f" ")
print(f"{EMOJI_OK} Proceso completado. Archivos generados: {num_archivos}")
print(f"{EMOJI_CLOCK} Tiempo total de ejecución: {minutos} Min {segundos} Seg")
print(f" ")
print(f" ")
print(f" ")