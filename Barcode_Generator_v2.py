import time
import os
import math
import pandas as pd
from barcode import Code128, EAN13
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

inicio = time.time()

# ---------- CONFIGURACIÓN ----------
FILAS_POR_ARCHIVO = 5000

GENERAR_CODE128 = False
OPTIONS_CODE128 = {
    "module_width": 0.2,
    "module_height": 8,
    "font_size": 8,
    "text_distance": 3.5,
    "quiet_zone": 4,
    "background": "white",
    "foreground": "black",
    "write_text": True,
    "dpi": 200
}
COLUMNA_CODE128 = "code128"
OPTIONS_CODE128_ANCHO_IMAGEN = 150
OPTIONS_CODE128_ALTO_IMAGEN = 45

GENERAR_EAN13 = True
OPTIONS_EAN13 = {
    "module_width": 0.2,
    "module_height": 8,
    "font_size": 6,
    "text_distance": 2.5,
    "quiet_zone": 4,
    "background": "white",
    "foreground": "black",
    "write_text": True,
    "dpi": 200
}
COLUMNA_EAN13 = "ean13"
OPTIONS_EAN13_ANCHO_IMAGEN = 150
OPTIONS_EAN13_ALTO_IMAGEN = 45

IS_WINDOWS = os.name == "nt"
EMOJI_OK = "✅" if not IS_WINDOWS else "[OK]"
EMOJI_CLOCK = "⏱️" if not IS_WINDOWS else "[TIEMPO]"
EMOJI_INFO = "ℹ️ " if not IS_WINDOWS else "[INFO]"
EMOJI_ALERT = "⚠️" if not IS_WINDOWS else "[ALERTA]"
EMOJI_ERROR = "❌" if not IS_WINDOWS else "[ERROR]"

# ---------- MENSAJE INICIAL ----------
print("\n\n")
print("******************************************************")
print("*** GENERADOR DE CÓDIGOS DE BARRAS PARA EXCEL v2 ***")
print("******************** By JoaquimCB ********************")
print("******************************************************")
print("\n\n")

# ---------- PEDIR AL USUARIO ----------
while True:
    directorio_entrada = input("Ingrese el directorio que contiene los archivos Excel: ").strip()
    if os.path.exists(directorio_entrada) and os.path.isdir(directorio_entrada):
        break
    else:
        print(f"{EMOJI_ERROR} Directorio no válido. Intente de nuevo.")

while True:
    COLUMNA_CODIGO = input("Ingrese el nombre de la columna que contiene los códigos de barras (Ej: cod_barras): ").strip()
    if COLUMNA_CODIGO:
        break
    else:
        print(f"{EMOJI_ERROR} Debe ingresar un nombre de columna válido.")

while True:
    try:
        FILAS_POR_ARCHIVO = int(input("Ingrese la cantidad de filas por archivo (Ej: 500): ").strip())
        if FILAS_POR_ARCHIVO > 0:
            break
        else:
            print(f"{EMOJI_ERROR} Debe ser un número mayor a 0.")
    except ValueError:
        print(f"{EMOJI_ERROR} Ingrese un número válido.")


# Carpeta de exportación
carpeta_exportacion = os.path.join(directorio_entrada, "Exportación")
os.makedirs(carpeta_exportacion, exist_ok=True)

# Listar todos los archivos Excel en el directorio
archivos_excel = [f for f in os.listdir(directorio_entrada) if f.endswith(".xlsx")]
if not archivos_excel:
    print(f"{EMOJI_ALERT} No se encontraron archivos Excel en el directorio indicado.")
    exit()

print(f"{EMOJI_INFO} Se encontraron {len(archivos_excel)} archivos Excel.\n")

# ---------- PROCESAR CADA EXCEL ----------
for archivo in archivos_excel:
    ruta_excel = os.path.join(directorio_entrada, archivo)
    nombre_base = os.path.splitext(archivo)[0]

    # Crear carpeta para este Excel dentro de Exportación
    carpeta_excel = os.path.join(carpeta_exportacion, nombre_base)
    os.makedirs(carpeta_excel, exist_ok=True)

    # Carpeta de barcodes
    carpeta_barcodes = os.path.join(carpeta_excel, "barcodes")
    os.makedirs(carpeta_barcodes, exist_ok=True)

    print(f"{EMOJI_INFO} Procesando {archivo} ...")

    # Leer Excel
    df = pd.read_excel(ruta_excel)
    if "cod_barras" not in df.columns:
        print(f"{EMOJI_ERROR} El archivo {archivo} no tiene la columna 'cod_barras', se saltará.")
        continue

    total_filas = len(df)
    num_archivos = (total_filas // FILAS_POR_ARCHIVO) + (1 if total_filas % FILAS_POR_ARCHIVO else 0)

    for bloque in range(num_archivos):
        inicio_bloque = bloque * FILAS_POR_ARCHIVO
        fin_bloque = inicio_bloque + FILAS_POR_ARCHIVO
        df_bloque = df.iloc[inicio_bloque:fin_bloque].copy()

        # Crear carpetas de imágenes para este bloque
        bloque_folder = os.path.join(carpeta_barcodes, f"bloque_{bloque+1}")
        os.makedirs(os.path.join(bloque_folder, "code128"), exist_ok=True)
        os.makedirs(os.path.join(bloque_folder, "ean13"), exist_ok=True)

        # Columnas vacías
        if GENERAR_CODE128:
            df_bloque[COLUMNA_CODE128] = ""
        if GENERAR_EAN13:
            df_bloque[COLUMNA_EAN13] = ""

        rutas_code128 = []
        rutas_ean13 = []

        # ---------- Generación de imágenes ----------
        print(f"{EMOJI_INFO} Generando imágenes para bloque {bloque+1} ...")
        for i, codigo in enumerate(df_bloque["cod_barras"]):
            if codigo is None or (isinstance(codigo, float) and math.isnan(codigo)):
                rutas_code128.append(None)
                rutas_ean13.append(None)
                continue

            try:
                codigo_str = str(int(codigo))
            except ValueError:
                print(f"Fila {i+2} tiene valor no numérico: {codigo}, se saltará")
                rutas_code128.append(None)
                rutas_ean13.append(None)
                continue

            if GENERAR_CODE128:
                ruta_c128 = os.path.join(bloque_folder, "code128", f"{codigo_str}_code128_{i}.png")
                Code128(codigo_str, writer=ImageWriter()).save(ruta_c128.replace(".png", ""), options=OPTIONS_CODE128)
                rutas_code128.append(ruta_c128)

            if GENERAR_EAN13:
                ean13_str = codigo_str.zfill(13)
                ruta_ean13 = os.path.join(bloque_folder, "ean13", f"{ean13_str}_ean13_{i}.png")
                EAN13(ean13_str, writer=ImageWriter()).save(ruta_ean13.replace(".png", ""), options=OPTIONS_EAN13)
                rutas_ean13.append(ruta_ean13)
        print(f"{EMOJI_OK} Imágenes generadas para bloque {bloque+1}")

        # Guardar Excel temporal
        temp_excel = os.path.join(bloque_folder, "temp.xlsx")
        df_bloque.to_excel(temp_excel, index=False)

        # Abrir Excel con openpyxl
        wb = load_workbook(temp_excel)
        ws = wb.active

        # Función para encontrar índice de columna por nombre
        def indice_columna(nombre_columna):
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == nombre_columna:
                    return idx
            return None

        idx_c128 = indice_columna(COLUMNA_CODE128) if GENERAR_CODE128 else None
        idx_ean13 = indice_columna(COLUMNA_EAN13) if GENERAR_EAN13 else None

        # ---------- Insertar imágenes ----------
        print(f"{EMOJI_INFO} Insertando imágenes en Excel bloque {bloque+1} ...")
        for i in range(2, ws.max_row + 1):
            if GENERAR_CODE128 and rutas_code128[i-2]:
                img_c128 = Image(rutas_code128[i-2])
                img_c128.width = OPTIONS_CODE128_ANCHO_IMAGEN
                img_c128.height = OPTIONS_CODE128_ALTO_IMAGEN
                celda_c128 = f"{ws.cell(row=1, column=idx_c128).column_letter}{i}"
                ws.add_image(img_c128, celda_c128)

            if GENERAR_EAN13 and rutas_ean13[i-2]:
                img_ean13 = Image(rutas_ean13[i-2])
                img_ean13.width = OPTIONS_EAN13_ANCHO_IMAGEN
                img_ean13.height = OPTIONS_EAN13_ALTO_IMAGEN
                celda_ean13 = f"{ws.cell(row=1, column=idx_ean13).column_letter}{i}"
                ws.add_image(img_ean13, celda_ean13)

            ws.row_dimensions[i].height = 50
        print(f"{EMOJI_OK} Imágenes insertadas en Excel bloque {bloque+1}")

        # ---------- Guardar Excel final ----------
        archivo_salida = os.path.join(carpeta_excel, f"{nombre_base}_barcodes_{bloque+1}.xlsx")
        wb.save(archivo_salida)
        os.remove(temp_excel)

        print(f"{EMOJI_OK} Generado {archivo_salida} con {len(df_bloque)} filas\n")

fin = time.time()
tiempo_total = fin - inicio
minutos = int(tiempo_total // 60)
segundos = int(tiempo_total % 60)
print("--------------------------------")
print(f"{EMOJI_OK} Proceso completado. Archivos procesados: {len(archivos_excel)}")
print(f"{EMOJI_CLOCK} Tiempo total de ejecución: {minutos} Min {segundos} Seg\n")
